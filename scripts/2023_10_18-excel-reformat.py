"""
Script to move a column in an Excel file to a new column location.

Requirements:
    - openpyxl
"""
import os

import openpyxl


def move_column_in_excel(file_path, src_col, dest_col):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Translate column names to numerical indices (assuming headers in row 2)
    col_indices = {cell.value: cell.column for cell in ws[2]}

    try:
        src_col_idx = col_indices[src_col]
        dest_col_idx = col_indices[dest_col]
    except KeyError:
        print(f"Columns {src_col} or {dest_col} not found in {file_path}. Skipping.")
        return

    ws.insert_cols(dest_col_idx + 1)

    for row in ws.iter_rows(min_col=src_col_idx, max_col=src_col_idx, min_row=1):
        for cell in row:
            dest_cell = ws.cell(row=cell.row, column=dest_col_idx + 1)
            dest_cell._style = cell._style
            dest_cell.value = cell.value

    ws.delete_cols(src_col_idx)

    wb.save(file_path)


if __name__ == "__main__":
    directory_path = ...
    src_col = ...
    dest_col = ...

    for filename in os.listdir(directory_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(directory_path, filename)
            try:
                move_column_in_excel(file_path, src_col, dest_col)
            except Exception as e:
                print(f"Error processing {file_path}: {e}")
